import os
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    Reference,
)
from openpyxl.formula.translate import Translator

from BatkaAI.actions.file_actions import find_file


# =========================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================================================


def _ensure_xlsx_name(filename):
    filename = str(filename).strip()

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    return filename


def _find_excel_file(filename):
    matches = find_file(filename)

    if not matches:
        return None

    for file_path in matches:
        if file_path.suffix.lower() == ".xlsx":
            return file_path

    return None


def _get_sheet(workbook, sheet_name=None):
    if sheet_name:
        sheet_name = str(sheet_name).strip()

        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

        if len(workbook.sheetnames) == 1:
            real_sheet = workbook.active

            print(
                f"Лист '{sheet_name}' не найден. "
                f"Использую единственный лист: {real_sheet.title}"
            )

            return real_sheet

        print(
            f"Лист '{sheet_name}' не найден. "
            f"Доступные листы: {', '.join(workbook.sheetnames)}"
        )

        return None

    return workbook.active


def _parse_color(color):
    if not color:
        return None

    color = (
        str(color)
        .replace("#", "")
        .strip()
        .upper()
    )

    if len(color) == 6:
        return color

    return None


def _coerce_excel_value(value):
    """
    Преобразует входные значения в нормальные Excel-типы.

    "150000" -> 150000
    "12.5" -> 12.5
    "-35" -> -35
    "=SUM(A1:A5)" остаётся формулой
    "Январь" остаётся текстом
    """

    if value is None:
        return None

    if isinstance(
        value,
        (int, float, bool),
    ):
        return value

    if not isinstance(value, str):
        return value

    text = value.strip()

    if not text:
        return ""

    # Формула Excel
    if text.startswith("="):
        return text

    # Убираем пробелы-разделители тысяч
    numeric_text = (
        text
        .replace("\u00A0", "")
        .replace(" ", "")
    )

    # Целое число
    if re.fullmatch(
        r"[+-]?\d+",
        numeric_text,
    ):
        try:
            return int(
                numeric_text
            )
        except ValueError:
            pass

    # Десятичное число через точку или запятую
    if re.fullmatch(
        r"[+-]?\d+[.,]\d+",
        numeric_text,
    ):
        try:
            return float(
                numeric_text.replace(
                    ",",
                    ".",
                )
            )
        except ValueError:
            pass

    return value


def _apply_cell_format(cell, data):
    font_name = data.get(
        "font_name"
    )

    font_size = data.get(
        "font_size"
    )

    bold = data.get(
        "bold"
    )

    italic = data.get(
        "italic"
    )

    underline = data.get(
        "underline"
    )

    font_color = _parse_color(
        data.get(
            "font_color"
        )
    )

    cell.font = Font(
        name=(
            font_name
            or cell.font.name
            or "Calibri"
        ),
        size=(
            float(font_size)
            if font_size
            else cell.font.sz
        ),
        bold=(
            bool(bold)
            if bold is not None
            else cell.font.bold
        ),
        italic=(
            bool(italic)
            if italic is not None
            else cell.font.italic
        ),
        underline=(
            "single"
            if underline
            else cell.font.underline
        ),
        color=(
            font_color
            if font_color
            else cell.font.color
        ),
    )

    fill_color = _parse_color(
        data.get(
            "fill_color"
        )
    )

    if fill_color:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color,
        )

    horizontal = data.get(
        "alignment"
    )

    if horizontal:
        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical="center",
            wrap_text=bool(
                data.get(
                    "wrap_text",
                    False,
                )
            ),
        )

    elif (
        data.get(
            "wrap_text"
        )
        is not None
    ):
        cell.alignment = Alignment(
            horizontal=(
                cell.alignment.horizontal
            ),
            vertical=(
                cell.alignment.vertical
            ),
            wrap_text=bool(
                data.get(
                    "wrap_text"
                )
            ),
        )

    number_format = data.get(
        "number_format"
    )

    if number_format:
        cell.number_format = (
            number_format
        )


def _autofit_sheet(sheet):
    for column_cells in sheet.columns:
        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            max_length = max(
                max_length,
                len(value),
            )

        width = min(
            max(
                max_length + 2,
                8,
            ),
            50,
        )

        sheet.column_dimensions[
            column_letter
        ].width = width


def _sanitize_table_name(name):
    name = str(
        name or "BatkaTable"
    ).strip()

    name = re.sub(
        r"[^A-Za-zА-Яа-яЁё0-9_]",
        "_",
        name,
    )

    if not name:
        name = "BatkaTable"

    if name[0].isdigit():
        name = (
            f"Table_{name}"
        )

    return name


def _table_name_exists(
    workbook,
    table_name,
):
    target = (
        table_name.lower()
    )

    for sheet in workbook.worksheets:
        for table in sheet.tables.values():
            if (
                table.displayName.lower()
                == target
            ):
                return True

    return False


def _make_unique_table_name(
    workbook,
    table_name,
):
    base = _sanitize_table_name(
        table_name
    )

    if not _table_name_exists(
        workbook,
        base,
    ):
        return base

    number = 2

    while True:
        candidate = (
            f"{base}_{number}"
        )

        if not _table_name_exists(
            workbook,
            candidate,
        ):
            return candidate

        number += 1


def _prepare_table_headers(
    sheet,
    cell_range,
):
    cells = sheet[
        cell_range
    ]

    if not cells:
        raise ValueError(
            "Диапазон таблицы пуст."
        )

    first_row = cells[0]

    used = set()

    for index, cell in enumerate(
        first_row,
        start=1,
    ):
        value = (
            ""
            if cell.value is None
            else str(
                cell.value
            ).strip()
        )

        if not value:
            value = (
                f"Столбец{index}"
            )

        original = value

        number = 2

        while (
            value.lower()
            in used
        ):
            value = (
                f"{original}_{number}"
            )

            number += 1

        used.add(
            value.lower()
        )

        cell.value = value


def _remove_same_autofilter(
    sheet,
    cell_range,
):
    if (
        sheet.auto_filter.ref
        and str(
            sheet.auto_filter.ref
        ).upper()
        == str(
            cell_range
        ).upper()
    ):
        sheet.auto_filter.ref = None


def _make_reference_from_range(
    sheet,
    cell_range,
):
    (
        min_col,
        min_row,
        max_col,
        max_row,
    ) = range_boundaries(
        cell_range
    )

    return Reference(
        sheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )


def _enable_excel_recalculation(
    workbook,
):
    """
    Просим Microsoft Excel полностью
    пересчитать формулы при открытии книги.
    """

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    except Exception:
        pass


# =========================================================
# EXCEL ENGINE
# =========================================================


def excel_edit(
    filename,
    operation,
    data=None,
):
    data = (
        data
        or {}
    )

    try:

        # =================================================
        # CREATE
        # =================================================

        if operation == "create":
            filename = (
                _ensure_xlsx_name(
                    filename
                )
            )

            file_path = (
                Path.home()
                / "Desktop"
                / filename
            )

            workbook = Workbook()

            sheet = (
                workbook.active
            )

            sheet_name = data.get(
                "sheet_name"
            )

            if sheet_name:
                sheet.title = str(
                    sheet_name
                )

            _enable_excel_recalculation(
                workbook
            )

            workbook.save(
                file_path
            )

            print(
                f"Excel-файл создан: "
                f"{file_path}"
            )

            return file_path

        # =================================================
        # FIND FILE
        # =================================================

        file_path = _find_excel_file(
            filename
        )

        if not file_path:
            print(
                f"Excel-файл не найден: "
                f"{filename}"
            )

            return None

        # =================================================
        # OPEN
        # =================================================

        if operation == "open":
            os.startfile(
                file_path
            )

            print(
                f"Excel-файл открыт: "
                f"{file_path}"
            )

            return file_path

        workbook = load_workbook(
            file_path
        )

        _enable_excel_recalculation(
            workbook
        )

        # =================================================
        # READ
        # =================================================

        if operation == "read":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            print(
                f"Excel: "
                f"{file_path}"
            )

            print(
                f"Лист: "
                f"{sheet.title}"
            )

            print(
                "-" * 60
            )

            for row in sheet.iter_rows(
                values_only=True
            ):
                values = [
                    (
                        ""
                        if value is None
                        else str(value)
                    )
                    for value in row
                ]

                if any(values):
                    print(
                        " | ".join(
                            values
                        )
                    )

            print(
                "-" * 60
            )

            return file_path

        # =================================================
        # LIST SHEETS
        # =================================================

        elif operation == "list_sheets":
            print(
                "Листы Excel:"
            )

            for index, name in enumerate(
                workbook.sheetnames,
                start=1,
            ):
                print(
                    f"{index}. {name}"
                )

            return file_path

        # =================================================
        # ADD SHEET
        # =================================================

        elif operation == "add_sheet":
            sheet_name = str(
                data.get(
                    "sheet_name",
                    "Новый лист",
                )
            )

            if (
                sheet_name
                in workbook.sheetnames
            ):
                print(
                    f"Лист уже существует: "
                    f"{sheet_name}"
                )

                return None

            workbook.create_sheet(
                title=sheet_name
            )

        # =================================================
        # RENAME SHEET
        # =================================================

        elif operation == "rename_sheet":
            old_name = data.get(
                "old_name"
            )

            new_name = data.get(
                "new_name"
            )

            if (
                old_name
                not in workbook.sheetnames
            ):
                print(
                    f"Лист не найден: "
                    f"{old_name}"
                )

                return None

            workbook[
                old_name
            ].title = new_name

        # =================================================
        # DELETE SHEET
        # =================================================

        elif operation == "delete_sheet":
            sheet_name = data.get(
                "sheet_name"
            )

            if (
                sheet_name
                not in workbook.sheetnames
            ):
                print(
                    f"Лист не найден: "
                    f"{sheet_name}"
                )

                return None

            if (
                len(
                    workbook.sheetnames
                )
                <= 1
            ):
                print(
                    "Нельзя удалить "
                    "единственный лист."
                )

                return None

            del workbook[
                sheet_name
            ]

        # =================================================
        # SET CELL
        # =================================================

        elif operation == "set_cell":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_address = str(
                data.get(
                    "cell",
                    "",
                )
            ).strip()

            if not cell_address:
                print(
                    "Не указана ячейка."
                )

                return None

            value = (
                _coerce_excel_value(
                    data.get(
                        "value"
                    )
                )
            )

            sheet[
                cell_address
            ] = value

        # =================================================
        # SET RANGE
        # =================================================

        elif operation == "set_range":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            start_cell = str(
                data.get(
                    "start_cell",
                    "A1",
                )
            )

            values = data.get(
                "values",
                [],
            )

            start = sheet[
                start_cell
            ]

            for (
                row_index,
                row_data,
            ) in enumerate(
                values
            ):
                for (
                    column_index,
                    value,
                ) in enumerate(
                    row_data
                ):
                    sheet.cell(
                        row=(
                            start.row
                            + row_index
                        ),
                        column=(
                            start.column
                            + column_index
                        ),
                        value=(
                            _coerce_excel_value(
                                value
                            )
                        ),
                    )

        # =================================================
        # APPEND ROW
        # =================================================

        elif operation == "append_row":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            values = [
                _coerce_excel_value(
                    value
                )
                for value in data.get(
                    "values",
                    [],
                )
            ]

            sheet.append(
                values
            )

        # =================================================
        # APPEND ROWS
        # =================================================

        elif operation == "append_rows":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            for row in data.get(
                "rows",
                [],
            ):
                clean_row = [
                    _coerce_excel_value(
                        value
                    )
                    for value in row
                ]

                sheet.append(
                    clean_row
                )

        # =================================================
        # SET FORMULA
        # =================================================

        elif operation == "set_formula":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_address = str(
                data.get(
                    "cell",
                    "",
                )
            ).strip()

            formula = str(
                data.get(
                    "formula",
                    "",
                )
            ).strip()

            if not cell_address:
                print(
                    "Не указана ячейка "
                    "формулы."
                )

                return None

            if not formula:
                print(
                    "Не указана формула."
                )

                return None

            if not formula.startswith(
                "="
            ):
                formula = (
                    "="
                    + formula
                )

            sheet[
                cell_address
            ] = formula

        # =================================================
        # FILL FORMULA DOWN
        # =================================================

        elif operation == "fill_formula_down":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            source_cell = str(
                data.get(
                    "source_cell",
                    "",
                )
            ).strip()

            end_cell = str(
                data.get(
                    "end_cell",
                    "",
                )
            ).strip()

            if (
                not source_cell
                or not end_cell
            ):
                print(
                    "Не указаны ячейки "
                    "для протягивания."
                )

                return None

            source = sheet[
                source_cell
            ]

            end = sheet[
                end_cell
            ]

            source_formula = (
                source.value
            )

            if (
                not isinstance(
                    source_formula,
                    str,
                )
                or not source_formula.startswith(
                    "="
                )
            ):
                print(
                    f"В {source_cell} "
                    f"нет формулы."
                )

                return None

            if (
                source.column
                != end.column
            ):
                print(
                    "Протягивание вниз "
                    "должно быть "
                    "в одном столбце."
                )

                return None

            for row_number in range(
                source.row + 1,
                end.row + 1,
            ):
                target = sheet.cell(
                    row=row_number,
                    column=source.column,
                )

                target.value = Translator(
                    source_formula,
                    origin=(
                        source.coordinate
                    ),
                ).translate_formula(
                    target.coordinate
                )

        # =================================================
        # MERGE
        # =================================================

        elif operation == "merge_cells":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            sheet.merge_cells(
                data.get(
                    "range"
                )
            )

        # =================================================
        # UNMERGE
        # =================================================

        elif operation == "unmerge_cells":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            sheet.unmerge_cells(
                data.get(
                    "range"
                )
            )

        # =================================================
        # FORMAT RANGE
        # =================================================

        elif operation == "format_range":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_range = data.get(
                "range"
            )

            if not cell_range:
                print(
                    "Не указан диапазон."
                )

                return None

            for row in sheet[
                cell_range
            ]:
                for cell in row:
                    _apply_cell_format(
                        cell,
                        data,
                    )

        # =================================================
        # FORMAT HEADER
        # =================================================

        elif operation == "format_header":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_range = data.get(
                "range",
                "A1:Z1",
            )

            fill_color = _parse_color(
                data.get(
                    "fill_color",
                    "1F4E78",
                )
            )

            font_color = _parse_color(
                data.get(
                    "font_color",
                    "FFFFFF",
                )
            )

            thin = Side(
                style="thin",
                color="D9E1F2",
            )

            for row in sheet[
                cell_range
            ]:
                for cell in row:
                    cell.font = Font(
                        bold=True,
                        color=font_color,
                    )

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=fill_color,
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                    cell.border = Border(
                        left=thin,
                        right=thin,
                        top=thin,
                        bottom=thin,
                    )

        # =================================================
        # COLUMN WIDTH
        # =================================================

        elif operation == "set_column_width":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            column = str(
                data.get(
                    "column",
                    "A",
                )
            ).upper()

            width = float(
                data.get(
                    "width",
                    15,
                )
            )

            sheet.column_dimensions[
                column
            ].width = width

        # =================================================
        # ROW HEIGHT
        # =================================================

        elif operation == "set_row_height":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            row_number = int(
                data.get(
                    "row",
                    1,
                )
            )

            height = float(
                data.get(
                    "height",
                    20,
                )
            )

            sheet.row_dimensions[
                row_number
            ].height = height

        # =================================================
        # AUTOFIT
        # =================================================

        elif operation == "autofit":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            _autofit_sheet(
                sheet
            )

        # =================================================
        # FREEZE PANES
        # =================================================

        elif operation == "freeze_panes":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            sheet.freeze_panes = (
                data.get(
                    "cell",
                    "A2",
                )
            )

        # =================================================
        # FILTER
        # =================================================

        elif operation == "add_filter":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_range = data.get(
                "range"
            )

            if not cell_range:
                print(
                    "Не указан диапазон "
                    "фильтра."
                )

                return None

            sheet.auto_filter.ref = (
                cell_range
            )

        # =================================================
        # SORT
        # =================================================

        elif operation == "sort":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_range = data.get(
                "range"
            )

            if not cell_range:
                print(
                    "Не указан диапазон "
                    "сортировки."
                )

                return None

            column_index = int(
                data.get(
                    "column_index",
                    1,
                )
            )

            descending = bool(
                data.get(
                    "descending",
                    False,
                )
            )

            rows = [
                [
                    cell.value
                    for cell in row
                ]
                for row in sheet[
                    cell_range
                ]
            ]

            if not rows:
                return None

            header = rows[0]
            body = rows[1:]

            def sort_key(row):
                value = row[
                    column_index - 1
                ]

                if value is None:
                    return (
                        1,
                        "",
                    )

                if isinstance(
                    value,
                    (int, float),
                ):
                    return (
                        0,
                        value,
                    )

                return (
                    0,
                    str(
                        value
                    ).lower(),
                )

            body.sort(
                key=sort_key,
                reverse=descending,
            )

            start = sheet[
                cell_range.split(
                    ":"
                )[0]
            ]

            for (
                row_offset,
                row_data,
            ) in enumerate(
                [header] + body
            ):
                for (
                    col_offset,
                    value,
                ) in enumerate(
                    row_data
                ):
                    sheet.cell(
                        row=(
                            start.row
                            + row_offset
                        ),
                        column=(
                            start.column
                            + col_offset
                        ),
                        value=value,
                    )

        # =================================================
        # FIND
        # =================================================

        elif operation == "find":
            query = str(
                data.get(
                    "query",
                    "",
                )
            ).strip().lower()

            if not query:
                print(
                    "Не указано значение "
                    "для поиска."
                )

                return None

            sheets = (
                workbook.worksheets
            )

            if data.get(
                "sheet_name"
            ):
                target_sheet = _get_sheet(
                    workbook,
                    data.get(
                        "sheet_name"
                    ),
                )

                sheets = (
                    [target_sheet]
                    if target_sheet
                    else []
                )

            results = []

            for sheet in sheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if (
                            cell.value
                            is not None
                            and query
                            in str(
                                cell.value
                            ).lower()
                        ):
                            results.append(
                                (
                                    sheet.title,
                                    cell.coordinate,
                                    cell.value,
                                )
                            )

            if not results:
                print(
                    f"В Excel ничего "
                    f"не найдено: {query}"
                )

            else:
                print(
                    f"Найдено совпадений: "
                    f"{len(results)}"
                )

                for (
                    found_sheet,
                    coordinate,
                    value,
                ) in results:
                    print(
                        f"{found_sheet}!"
                        f"{coordinate}: "
                        f"{value}"
                    )

            return file_path

        # =================================================
        # REPLACE
        # =================================================

        elif operation == "replace":
            old_text = str(
                data.get(
                    "old_text",
                    "",
                )
            )

            new_text = str(
                data.get(
                    "new_text",
                    "",
                )
            )

            if not old_text:
                print(
                    "Не указан текст "
                    "для замены."
                )

                return None

            replacements = 0

            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if (
                            isinstance(
                                cell.value,
                                str,
                            )
                            and old_text
                            in cell.value
                        ):
                            cell.value = (
                                cell.value.replace(
                                    old_text,
                                    new_text,
                                )
                            )

                            replacements += 1

            print(
                f"Выполнено замен: "
                f"{replacements}"
            )

        # =================================================
        # CREATE TABLE
        # =================================================

        elif operation == "create_table":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_range = str(
                data.get(
                    "range",
                    "",
                )
            ).strip()

            if not cell_range:
                print(
                    "Не указан диапазон "
                    "таблицы."
                )

                return None

            cells = sheet[
                cell_range
            ]

            if len(cells) < 2:
                print(
                    "Для Excel-таблицы "
                    "нужна строка заголовков "
                    "и данные."
                )

                return None

            _prepare_table_headers(
                sheet,
                cell_range,
            )

            for existing_table in (
                sheet.tables.values()
            ):
                if (
                    str(
                        existing_table.ref
                    ).upper()
                    == cell_range.upper()
                ):
                    print(
                        "На этом диапазоне "
                        "уже существует таблица "
                        f"{existing_table.displayName}"
                    )

                    return file_path

            table_name = (
                _make_unique_table_name(
                    workbook,
                    data.get(
                        "table_name",
                        "BatkaTable",
                    ),
                )
            )

            _remove_same_autofilter(
                sheet,
                cell_range,
            )

            table = Table(
                displayName=table_name,
                ref=cell_range,
            )

            table.tableStyleInfo = (
                TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
            )

            sheet.add_table(
                table
            )

            print(
                f"Excel-таблица создана: "
                f"{table_name} "
                f"({sheet.title}!"
                f"{cell_range})"
            )

        # =================================================
        # CREATE CHART
        # =================================================

        elif operation == "create_chart":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            chart_type = str(
                data.get(
                    "chart_type",
                    "bar",
                )
            ).lower()

            if chart_type == "line":
                chart = LineChart()

            elif chart_type == "pie":
                chart = PieChart()

            else:
                chart = BarChart()

                # Для вертикальных столбцов
                chart.type = "col"
                chart.style = 10

            data_range = data.get(
                "data_range"
            )

            category_range = data.get(
                "category_range"
            )

            if not data_range:
                print(
                    "Не указан диапазон данных "
                    "для диаграммы."
                )

                return None

            if not category_range:
                print(
                    "Не указан диапазон категорий "
                    "для диаграммы."
                )

                return None

            data_ref = (
                _make_reference_from_range(
                    sheet,
                    data_range,
                )
            )

            category_ref = (
                _make_reference_from_range(
                    sheet,
                    category_range,
                )
            )

            chart.add_data(
                data_ref,
                titles_from_data=True,
                from_rows=False,
            )

            chart.set_categories(
                category_ref
            )

            chart_title = str(
                data.get(
                    "title",
                    "Диаграмма",
                )
            )

            chart.title = (
                chart_title
            )

            chart.height = float(
                data.get(
                    "height",
                    8,
                )
            )

            chart.width = float(
                data.get(
                    "width",
                    14,
                )
            )

            if chart_type != "pie":
                chart.y_axis.title = (
                    data.get(
                        "y_axis_title",
                        ""
                    )
                )

                chart.x_axis.title = (
                    data.get(
                        "x_axis_title",
                        ""
                    )
                )

            position = data.get(
                "position",
                "F2",
            )

            sheet.add_chart(
                chart,
                position,
            )

            print(
                f"Диаграмма создана: "
                f"{chart_title}"
            )

        # =================================================
        # CLEAR RANGE
        # =================================================

        elif operation == "clear_range":
            sheet = _get_sheet(
                workbook,
                data.get(
                    "sheet_name"
                ),
            )

            if not sheet:
                return None

            cell_range = data.get(
                "range"
            )

            if not cell_range:
                print(
                    "Не указан диапазон."
                )

                return None

            for row in sheet[
                cell_range
            ]:
                for cell in row:
                    cell.value = None

        # =================================================
        # SAVE AS
        # =================================================

        elif operation == "save_as":
            new_filename = str(
                data.get(
                    "new_filename",
                    "",
                )
            ).strip()

            if not new_filename:
                print(
                    "Не указано новое имя "
                    "Excel-файла."
                )

                return None

            new_filename = (
                _ensure_xlsx_name(
                    new_filename
                )
            )

            new_path = (
                Path.home()
                / "Desktop"
                / new_filename
            )

            _enable_excel_recalculation(
                workbook
            )

            workbook.save(
                new_path
            )

            print(
                f"Excel сохранён как: "
                f"{new_path}"
            )

            return new_path

        # =================================================
        # UNKNOWN
        # =================================================

        else:
            print(
                f"Неизвестная операция "
                f"Excel: {operation}"
            )

            return None

        # =================================================
        # SAVE
        # =================================================

        _enable_excel_recalculation(
            workbook
        )

        workbook.save(
            file_path
        )

        print(
            f"Excel-файл изменён: "
            f"{file_path}"
        )

        return file_path

    except PermissionError as e:
        print(
            "Не удалось изменить Excel-файл. "
            "Закрой его в Microsoft Excel "
            "и повтори команду."
        )

        print(
            f"Техническая ошибка: "
            f"{type(e).__name__}: "
            f"{repr(e)}"
        )

        return None

    except Exception as e:
        print(
            f"Ошибка Excel Engine: "
            f"{type(e).__name__}: "
            f"{repr(e)}"
        )

        return None